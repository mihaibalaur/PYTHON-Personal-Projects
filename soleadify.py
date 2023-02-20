import pandas as pd
from openpyxl.utils.cell import get_column_letter

# print("-" * 100) - used to delimiter the "print" function's outputs.
# FACEBOOK DATASET - READING & CLEANING
colnames_facebook = ["domain",
                     "adress",
                     "categories",
                     "city",
                     "country_code",
                     "country_name",
                     "description",
                     "email",
                     "link",
                     "name",
                     "page_type",
                     "phone",
                     "phone_country_code",
                     "region_code",
                     "region_name",
                     "zip_code"]

# Due to the fact that there are values on columns that don't have a header,
# I created e FOR LOOP in order to append the already existing column heads +
# letters/ letters combinations as Excel provides (A, AB, BC) to read the entire dataset.

# x - the number of rows that contain values
x = 231
columns = [get_column_letter(i + 1) for i in range(x)]
for i in columns:
    colnames_facebook.append(i)

print("The columns of the Facebook Dataset are:")
print(colnames_facebook)
print("-" * 100)

facebook = pd.read_csv("facebook_dataset.csv",
                       names=colnames_facebook,
                       low_memory=False,
                       encoding='utf-8')

# Convert column 'phone' to numeric
# errors='coerce' - convert non-numeric values to NaN Values.
facebook['phone'] = pd.to_numeric(facebook['phone'], errors='coerce')

# Drop rows with non-numeric values in column 'phone'.
facebook = facebook.loc[~facebook['phone'].isna()]

# Converting the "phone" column to float32
facebook["phone"] = facebook["phone"].astype('float32')

# Option that allows displaying all the columns, using the "head()" function.
pd.options.display.max_columns = None

# Display the first 5 rows of the DataFrame.
print(facebook.head(3))
print("-" * 100)

# Drop the first row of values because it contains the initial header
# and no relevant data for the columns.
facebook.drop(index=facebook.index[0],
              axis=0,
              inplace=True)

# Display information about the Dataframe.
print(facebook.info())
print("-" * 100)

# GOOGLE DATASET - READING & CLEANING
colnames_google = ["adress",
                   "category",
                   "city",
                   "country_code",
                   "country_name",
                   "name",
                   "phone",
                   "phone_country_code",
                   "raw_adress",
                   "raw_phone",
                   "region_code",
                   "region_name",
                   "text",
                   "zip_code",
                   "domain"]

y = 15
columns = [get_column_letter(i + 1) for i in range(y)]
for i in columns:
    colnames_google.append(i)

print("The columns of the Google Dataset are:")
print(colnames_google)
print("-" * 100)

google = pd.read_csv("google_dataset.csv",
                     names=colnames_google,
                     low_memory=False,
                     encoding='utf-8')

# Only selecting the rows that have a numeric value on "phone" column.
# \d+ matches one or more characters that are not digits.
google['phone'] = google['phone'].str.extract(pat="(\d+)",
                                              expand=False)

# Converting the "phone" column from object to float32
google["phone"] = google["phone"].astype('float32')

print(google.head(3))
print("-" * 100)

google.drop(index=google.index[0],
            axis=0,
            inplace=True)

print(google.info())
print("-" * 100)

# WEBSITE DATASET - READING & CLEANING

colnames_website_names = ["root_domain",
                          "domain_suffix",
                          "language",
                          "legal_name",
                          "main_city",
                          "main_country",
                          "main_region",
                          "phone",
                          "site_name",
                          "tld",
                          "s_category"]

colnames_website_letters = []

x = 19
columns = [get_column_letter(i + 1) for i in range(x)]
for i in columns:
    colnames_website_letters.append(i)

website = pd.read_csv("website_dataset.csv",
                      names=colnames_website_letters,
                      low_memory=False,
                      encoding='utf-8')

print(website.head(3))
print("-" * 100)

# Select only the first series of the DataFrame.
first_column = website.iloc[:, 0]
print(first_column)
print("-" * 100)

# All the values are on the first column and must be separated by ";".
# website[colnames_website_names] - the columns where to put expand the values on.
split = website[colnames_website_names] = first_column.str.split(";", expand=True)
print(split)
print("-" * 100)

# The column was separated into columns, that were appended to the end of the DataFrame
print(website.head(3))
print("-" * 100)

# Drop the first row of the Dataframe.
drop = website.drop(index=website.index[0],
                    axis=0,
                    inplace=True)

# Change the positions of columns.
website = website.reindex(columns=colnames_website_names+colnames_website_letters)
print(website.head(3))
print("-" * 100)

# Drop the column "A"
website = website.drop('A', axis=1)
print(website.head(3))
print("-" * 100)

# Convert column 'phone' to numeric
website['phone'] = pd.to_numeric(website['phone'], errors='coerce')

# Drop rows with non-numeric values in column 'phone',
website = website.loc[~website['phone'].isna()]

# Convert the "phone" column to float32,
website["phone"] = website["phone"].astype('float32')

# JOIN -
# Create subsets of the Dataframe
# The subsets include all the necessary columns, that will be later used in joining.

facebook_subset = facebook[["name",
                            "categories",
                            "phone",
                            "city",
                            "region_name",
                            "country_name"]]

google_subset = google[["name",
                        "category",
                        "phone",
                        "city",
                        "region_name",
                        "country_name"]]

website_subset = website[["legal_name",
                          "s_category",
                          "phone",
                          "main_city",
                          "main_region",
                          "main_country"]]

# Deactivate the chained assignment mode while working on slices of the dataframe.
pd.options.mode.chained_assignment = None

# Rename the columns accordingly.
facebook_subset.rename(columns={"categories": "category_FB",
                                "city": "city_FB",
                                "name": "name_FB",
                                "phone": "phone_FB",
                                "region_name": "region_name_FB"}, inplace=True)

google_subset.rename(columns={"category": "category_GG",
                              "city": "city_GG",
                              "name": "name_GG",
                              "phone": "phone_GG",
                              "region_name": "region_name_GG"}, inplace=True)

website_subset.rename(columns={"s_category": "category_WS",
                               "main_city": "city_WS",
                               "legal_name": "name_WS",
                               "phone": "phone_WS",
                               "main_country": "country_name",
                               "main_region": "region_name_WS"}, inplace=True)
# Information Check.
print(facebook_subset.info())
print("-" * 100)

print(google_subset.info())
print("-" * 100)

print(website_subset.info())
print("-" * 100)

# Drop the NaN Values from the subsets.
facebook_subset = facebook_subset.dropna(axis=0)

google_subset = google_subset.dropna(axis=0)

website_subset = website_subset.dropna(axis=0)

# Check again to see if any changes actually happened.
print(facebook_subset.info())
print("-" * 100)

print(google_subset.info())
print("-" * 100)

print(website_subset.info())
print("-" * 100)

# Use only the first 500 rows of each subset, due to low memory capacity.
facebook_part = facebook_subset[:500]

google_part = google_subset[:500]

website_part = website_subset[:500]

# JOIN
# I attempted to join the datasets, using 2 different columns separately.

# A.1) First JOIN - between FACEBOOK & GOOGLE on "index".
join_index1 = pd.merge(facebook_part,
                       google_part,
                       left_index=True,
                       right_index=True)

# A.2) Second JOIN - between JOIN_INDEX1 & WEBSITE on "index"
join_index2 = pd.merge(join_index1,
                       website_part,
                       left_index=True,
                       right_index=True)

print(join_index2.head(3))

# B.1) First JOIN - between FACEBOOK & GOOGLE on "country_name";
join1 = pd.merge(facebook_part,
                 google_part,
                 how='left',
                 on='country_name')

# B.2) Second JOIN - between JOIN1 & WEBSITE on "country_name";
join2 = pd.merge(join1,
                 website_part,
                 how='inner',
                 on='country_name')
print(join2.head(3))


# In order to join the datasets, I used the "country_name" column because:

# 1. each dataset has an already existing column filled with country values,
# therefore there is no need to create a new one to join on.

# 2. it contains values on a larger and more general scale,
# making the join more widely specific.
